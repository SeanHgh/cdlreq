"""Export functionality for requirements and specifications"""

from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from .models import Requirement, Specification


class TraceabilityMatrixExporter:
    """Export requirements and specifications to Excel traceability matrix"""
    
    def __init__(self, requirements: List[Requirement], specifications: List[Specification]):
        self.requirements = requirements
        self.specifications = specifications
        self.req_map = {req.id: req for req in requirements}
        self.spec_map = {spec.id: spec for spec in specifications}
    
    def export_to_excel(self, output_path: Path) -> None:
        """Export traceability matrix to Excel file"""
        workbook = openpyxl.Workbook()
        
        # Remove default sheet and create our sheets
        workbook.remove(workbook.active)
        
        # Create sheets
        self._create_requirements_sheet(workbook)
        self._create_specifications_sheet(workbook)
        self._create_traceability_matrix_sheet(workbook)
        self._create_summary_sheet(workbook)
        
        # Save workbook
        workbook.save(output_path)
    
    def _create_requirements_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create requirements sheet"""
        ws = workbook.create_sheet("Requirements", 0)
        
        # Headers
        headers = ["ID", "Title", "Description", "Type", "Tags", "Acceptance Criteria", "Source", "Rationale"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        for row, req in enumerate(self.requirements, 2):
            ws.cell(row=row, column=1, value=req.id)
            ws.cell(row=row, column=2, value=req.title)
            ws.cell(row=row, column=3, value=req.description)
            ws.cell(row=row, column=4, value=req.type)
            ws.cell(row=row, column=5, value=", ".join(req.tags))
            ws.cell(row=row, column=6, value="\n".join(req.acceptance_criteria))
            ws.cell(row=row, column=7, value=req.source or "")
            ws.cell(row=row, column=8, value=req.rationale or "")
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_specifications_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create specifications sheet"""
        ws = workbook.create_sheet("Specifications", 1)
        
        # Headers
        headers = ["ID", "Title", "Description", "Related Requirements", "Implementation Unit", 
                  "Unit Test", "Test Criteria", "Design Notes", "Dependencies"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        for row, spec in enumerate(self.specifications, 2):
            ws.cell(row=row, column=1, value=spec.id)
            ws.cell(row=row, column=2, value=spec.title)
            ws.cell(row=row, column=3, value=spec.description)
            ws.cell(row=row, column=4, value=", ".join(spec.related_requirements))
            ws.cell(row=row, column=5, value=spec.implementation_unit)
            ws.cell(row=row, column=6, value=spec.unit_test)
            ws.cell(row=row, column=7, value="\n".join(spec.test_criteria))
            ws.cell(row=row, column=8, value=spec.design_notes or "")
            ws.cell(row=row, column=9, value=", ".join(spec.dependencies))
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_traceability_matrix_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create traceability matrix sheet"""
        ws = workbook.create_sheet("Traceability Matrix", 2)
        
        # Title
        ws.cell(row=1, column=1, value="Requirements to Specifications Traceability Matrix")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        
        # Headers
        ws.cell(row=3, column=1, value="Requirement ID")
        ws.cell(row=3, column=2, value="Requirement Title")
        ws.cell(row=3, column=3, value="Requirement Type")
        ws.cell(row=3, column=4, value="Specification ID")
        ws.cell(row=3, column=5, value="Specification Title")
        ws.cell(row=3, column=6, value="Implementation Unit")
        ws.cell(row=3, column=7, value="Unit Test")
        ws.cell(row=3, column=8, value="Design Notes")
        
        # Style header row
        for col in range(1, 9):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Build requirement to specifications mapping
        req_to_specs = {}
        for spec in self.specifications:
            for req_id in spec.related_requirements:
                if req_id not in req_to_specs:
                    req_to_specs[req_id] = []
                req_to_specs[req_id].append(spec)
        
        # Data rows
        current_row = 4
        for req in self.requirements:
            related_specs = req_to_specs.get(req.id, [])
            
            if not related_specs:
                # Requirement with no specifications
                ws.cell(row=current_row, column=1, value=req.id)
                ws.cell(row=current_row, column=2, value=req.title)
                ws.cell(row=current_row, column=3, value=req.type)
                ws.cell(row=current_row, column=4, value="No specifications")
                
                # Style untraced requirement
                for col in range(1, 9):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                current_row += 1
            else:
                # Requirement with specifications
                start_row = current_row
                
                for i, spec in enumerate(related_specs):
                    if i == 0:
                        # First specification - add requirement info
                        ws.cell(row=current_row, column=1, value=req.id)
                        ws.cell(row=current_row, column=2, value=req.title)
                        ws.cell(row=current_row, column=3, value=req.type)
                    
                    # Add specification info
                    ws.cell(row=current_row, column=4, value=spec.id)
                    ws.cell(row=current_row, column=5, value=spec.title)
                    ws.cell(row=current_row, column=6, value=spec.implementation_unit)
                    ws.cell(row=current_row, column=7, value=spec.unit_test)
                    ws.cell(row=current_row, column=8, value=spec.design_notes or "")
                    
                    # Style traced requirement-specification pair
                    for col in range(1, 9):
                        cell = ws.cell(row=current_row, column=col)
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    
                    current_row += 1
                
                # Merge requirement cells if multiple specifications
                if len(related_specs) > 1:
                    end_row = current_row - 1
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                    ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                    
                    # Center align merged cells
                    for col in range(1, 4):
                        cell = ws.cell(row=start_row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add borders
        self._add_borders(ws, 3, 1, current_row - 1, 8)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Create summary sheet"""
        ws = workbook.create_sheet("Summary", 3)
        
        # Title
        ws.cell(row=1, column=1, value="Traceability Matrix Summary")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Export info
        ws.cell(row=3, column=1, value="Export Date:")
        ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Statistics
        ws.cell(row=5, column=1, value="Statistics:")
        ws.cell(row=5, column=1).font = Font(bold=True)
        
        ws.cell(row=6, column=1, value="Total Requirements:")
        ws.cell(row=6, column=2, value=len(self.requirements))
        
        ws.cell(row=7, column=1, value="Total Specifications:")
        ws.cell(row=7, column=2, value=len(self.specifications))
        
        # Requirements by type
        req_by_type = {}
        for req in self.requirements:
            req_by_type[req.type] = req_by_type.get(req.type, 0) + 1
        
        ws.cell(row=9, column=1, value="Requirements by Type:")
        ws.cell(row=9, column=1).font = Font(bold=True)
        
        for i, (req_type, count) in enumerate(req_by_type.items(), 10):
            ws.cell(row=i, column=1, value=f"{req_type.title()}:")
            ws.cell(row=i, column=2, value=count)
        
        # Traceability analysis
        traced_reqs = set()
        for spec in self.specifications:
            traced_reqs.update(spec.related_requirements)
        
        untraced_reqs = set(req.id for req in self.requirements) - traced_reqs
        
        ws.cell(row=len(req_by_type) + 12, column=1, value="Traceability Analysis:")
        ws.cell(row=len(req_by_type) + 12, column=1).font = Font(bold=True)
        
        ws.cell(row=len(req_by_type) + 13, column=1, value="Requirements with Specifications:")
        ws.cell(row=len(req_by_type) + 13, column=2, value=len(traced_reqs))
        
        ws.cell(row=len(req_by_type) + 14, column=1, value="Requirements without Specifications:")
        ws.cell(row=len(req_by_type) + 14, column=2, value=len(untraced_reqs))
        
        if untraced_reqs:
            ws.cell(row=len(req_by_type) + 16, column=1, value="Untraced Requirements:")
            ws.cell(row=len(req_by_type) + 16, column=1).font = Font(bold=True)
            for i, req_id in enumerate(sorted(untraced_reqs), len(req_by_type) + 17):
                ws.cell(row=i, column=1, value=req_id)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """Auto-adjust column widths based on content"""
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter
        
        # Get the maximum column index to iterate through
        max_column = ws.max_column
        
        for col_idx in range(1, max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Check all cells in this column
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with some padding, but cap at reasonable maximum
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width > 0:  # Only set width if we found content
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_borders(self, ws: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
        """Add borders to a range of cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border